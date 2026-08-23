"""
Google Drive and Sheets connector v2.
Fresh sheet. On first run processes ALL files oldest→newest.
On subsequent runs skips already-processed files.
Uses GOOGLE_ACCESS_TOKEN from Replit Secrets.
"""
import io
import logging
import os
import threading
import time
import requests
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

SPREADSHEET_FILE_ID = os.getenv("GOOGLE_SHEETS_ID", "151jig9_3v-__dHfk7TksitJYONDMOLQV")

# ── In-process raw-bytes cache ─────────────────────────────────────────────────
# Avoids re-downloading the full xlsx on every write (Progress Report saves,
# deletes, appends). TTL matches the Streamlit cache (3600 s); _save_and_upload()
# refreshes it immediately after each upload so reads after writes see fresh data.
_RAW_CACHE: dict = {"data": None, "ts": 0.0}
_RAW_CACHE_TTL = 3600  # seconds
_WORKBOOK_LOCK = threading.RLock()


def _serialized_workbook_write(func):
    """Serialize writes in this app process and always start from fresh bytes."""
    def wrapped(*args, **kwargs):
        with _WORKBOOK_LOCK:
            _invalidate_raw_cache()
            return func(*args, **kwargs)
    wrapped.__name__ = func.__name__
    wrapped.__doc__ = func.__doc__
    return wrapped


def _get_master_bytes() -> bytes:
    """Return master xlsx bytes, from cache if ≤30 s old."""
    if _RAW_CACHE["data"] is None or (time.monotonic() - _RAW_CACHE["ts"]) > _RAW_CACHE_TTL:
        _RAW_CACHE["data"] = download_drive_file(SPREADSHEET_FILE_ID)
        _RAW_CACHE["ts"]   = time.monotonic()
    return _RAW_CACHE["data"]


def _invalidate_raw_cache() -> None:
    _RAW_CACHE["data"] = None
    _RAW_CACHE["ts"]   = 0.0


def warm_bytes_cache() -> None:
    """Download the master workbook bytes into the in-process cache.
    Safe to call from a background thread — no Streamlit dependency."""
    try:
        _get_master_bytes()
    except Exception:
        pass
SHEET_TAB_NAME      = "Opportunities"
SHORTLISTED_TAB_NAME = "Shortlisted"
URGENT_TAB_NAME     = "Urgent"

OUTPUT_COLUMNS = [
    "Solicitation Number",
    "Title",
    "Agency",
    "Solicitation Date",
    "Due Date",
    "Opportunity Type",
    "Normalized Set Aside",
    "UiLink",
    "Progress Report",
    "Award Status",
    "Assigned To",
]

SHORTLISTED_COLUMNS = OUTPUT_COLUMNS + [
    "Shortlisted By",
    "Shortlisted At",
    "Last Updated By",
    "Last Updated At",
    "Team Notes",
]

PROGRESS_OPTIONS = (
    "Bid Submitted,Bid InProgress,Sub Contractor Inquiry,"
    "Bid Past Due Date,Bid Quote Requested"
)


# ─────────────────────────────────────────────
# Auth
# ─────────────────────────────────────────────

def _token() -> str:
    from token_store import get_token
    return get_token("google-drive")


def _headers(extra: dict = None) -> dict:
    h = {"Authorization": f"Bearer {_token()}"}
    if extra:
        h.update(extra)
    return h


# ─────────────────────────────────────────────
# Drive helpers
# ─────────────────────────────────────────────

def list_drive_files(folder_id: str) -> list:
    """List all CSV/xlsx files, oldest first."""
    resp = requests.get(
        "https://www.googleapis.com/drive/v3/files",
        params={
            "q": f"'{folder_id}' in parents and trashed=false",
            "fields": "files(id,name,mimeType,modifiedTime,createdTime)",
            "orderBy": "createdTime asc",   # oldest first
            "pageSize": 200,
        },
        headers=_headers(),
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json().get("files", [])


def download_drive_file(file_id: str) -> bytes:
    resp = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"alt": "media"},
        headers=_headers(),
        timeout=60,
    )
    resp.raise_for_status()
    return resp.content


def upload_drive_file(file_id: str, content: bytes, mime_type: str) -> dict:
    resp = requests.patch(
        f"https://www.googleapis.com/upload/drive/v3/files/{file_id}",
        params={"uploadType": "media"},
        headers=_headers({"Content-Type": mime_type}),
        data=content,
        timeout=120,
    )
    resp.raise_for_status()
    return resp.json()


# ─────────────────────────────────────────────
# Sheet setup
# ─────────────────────────────────────────────

def _open_master() -> tuple:
    """Load master xlsx for writing (uses byte cache). Returns (wb, ws, raw_bytes)."""
    raw = _get_master_bytes()
    wb  = openpyxl.load_workbook(io.BytesIO(raw))

    # Create fresh tab if it doesn't exist
    if SHEET_TAB_NAME not in wb.sheetnames:
        log.info(f"Creating new sheet tab: '{SHEET_TAB_NAME}'")
        ws = wb.create_sheet(SHEET_TAB_NAME)
        _write_headers(ws)
    else:
        ws = wb[SHEET_TAB_NAME]
        # If sheet exists but has no headers, write them
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            _write_headers(ws)

    return wb, ws, raw


def _open_master_readonly():
    """Load master xlsx read-only (uses byte cache, no styles)."""
    raw = _get_master_bytes()
    wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws  = wb[SHEET_TAB_NAME] if SHEET_TAB_NAME in wb.sheetnames else wb.active
    return wb, ws


def _write_headers(ws):
    """Write bold headers to row 1."""
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_fill and header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    widths = [20, 40, 35, 18, 18, 20, 35, 40, 20, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    log.info(f"Headers written to sheet '{ws.title}'")


def _ensure_tab(wb, tab_name: str, columns: list):
    """Return a worksheet, creating and formatting it when necessary."""
    if tab_name in wb.sheetnames:
        ws = wb[tab_name]
    else:
        ws = wb.create_sheet(tab_name)

    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"
    return ws


def _save_and_upload(wb) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    data = buf.getvalue()
    upload_drive_file(
        SPREADSHEET_FILE_ID,
        data,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Refresh cache with what we just uploaded so the next read is instant
    _RAW_CACHE["data"] = data
    _RAW_CACHE["ts"]   = time.monotonic()
    log.info("✅ Master sheet uploaded successfully")


# ─────────────────────────────────────────────
# Dedup keys
# ─────────────────────────────────────────────

def get_existing_dedup_keys() -> tuple:
    """Return keys from both intake and shortlisted tabs to prevent re-imports."""
    try:
        raw = _get_master_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sol_nums, uilinks = set(), set()

        for tab_name in (SHEET_TAB_NAME, SHORTLISTED_TAB_NAME):
            if tab_name not in wb.sheetnames:
                continue
            ws = wb[tab_name]
            headers = [str(v or "").strip() for v in next(ws.iter_rows(values_only=True), [])]
            sol_col = headers.index("Solicitation Number") if "Solicitation Number" in headers else None
            link_col = headers.index("UiLink") if "UiLink" in headers else None
            for values in ws.iter_rows(min_row=2, values_only=True):
                if sol_col is not None and sol_col < len(values) and values[sol_col]:
                    sol_nums.add(str(values[sol_col]).strip())
                if link_col is not None and link_col < len(values) and values[link_col]:
                    uilinks.add(str(values[link_col]).strip())
        wb.close()

        log.info(f"Existing dedup keys: {len(sol_nums)} sol#, {len(uilinks)} links")
        return sol_nums, uilinks
    except Exception as e:
        log.warning(f"Could not load dedup keys: {e}")
        return set(), set()


# Legacy wrapper
def get_existing_solicitation_numbers() -> set:
    s, _ = get_existing_dedup_keys()
    return s


# ─────────────────────────────────────────────
# Append rows
# ─────────────────────────────────────────────

@_serialized_workbook_write
def append_rows_to_xlsx(rows: list, output_columns: list) -> int:
    """Append rows to master sheet. Returns count appended."""
    if not rows:
        return 0

    wb, ws, _ = _open_master()

    # Build header map
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            header_map[str(v).strip()] = c

    # Find true next empty row
    next_row = 2
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(r, c).value for c in range(1, len(OUTPUT_COLUMNS) + 2)):
            next_row = r + 1
            break

    log.info(f"Writing {len(rows)} rows starting at row {next_row}")

    for row_dict in rows:
        for col_name in output_columns:
            col_idx = header_map.get(col_name.strip())
            if col_idx is None:
                continue
            val = row_dict.get(col_name, "")
            ws.cell(row=next_row, column=col_idx, value=val)
        next_row += 1

    # Refresh dropdown
    _apply_dropdown(ws, header_map)

    _save_and_upload(wb)
    return len(rows)


@_serialized_workbook_write
def apply_progress_dropdown() -> None:
    """Re-apply Progress Report dropdown. Safe to call anytime."""
    wb, ws, _ = _open_master()
    header_map = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    _apply_dropdown(ws, header_map)
    _save_and_upload(wb)


def _apply_dropdown(ws, header_map: dict):
    col_idx = header_map.get("Progress Report")
    if col_idx is None:
        return
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    last_row   = max(ws.max_row, 1000)
    dv_range   = f"{col_letter}2:{col_letter}{last_row}"

    ws.data_validations.dataValidation = [
        v for v in ws.data_validations.dataValidation
        if col_letter not in str(v.sqref)
    ]
    dv = DataValidation(
        type="list",
        formula1=f'"{PROGRESS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid option",
        error="Please choose a value from the dropdown list.",
    )
    dv.sqref = dv_range
    ws.add_data_validation(dv)


# ─────────────────────────────────────────────
# Read sheet data (for Solicitations page)
# ─────────────────────────────────────────────

def read_master_sheet():
    """Download master sheet and return data as a pandas DataFrame."""
    import pandas as pd

    wb, ws = _open_master_readonly()
    try:
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not all_rows:
        return pd.DataFrame()

    headers = [str(v) if v is not None else f"Col{i+1}" for i, v in enumerate(all_rows[0])]
    data = [list(row) for row in all_rows[1:] if any(v is not None for v in row)]

    df = pd.DataFrame(data, columns=headers)
    if "Progress Report" in df.columns:
        df["Progress Report"] = df["Progress Report"].fillna("").astype(str).replace("None", "")
    # Rename legacy "Award Date" header → "Award Status" transparently
    if "Award Date" in df.columns and "Award Status" not in df.columns:
        df = df.rename(columns={"Award Date": "Award Status"})
    if "Award Status" in df.columns:
        df["Award Status"] = df["Award Status"].fillna("").astype(str).replace("None", "")
    # Inject Assigned To so the UI always renders the column even before
    # the first assignment is written back to the sheet.
    if "Assigned To" not in df.columns:
        df["Assigned To"] = ""
    return df


def read_shortlisted_sheet():
    """Read the active-work queue from the Shortlisted worksheet."""
    import pandas as pd

    raw = _get_master_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if SHORTLISTED_TAB_NAME not in wb.sheetnames:
            return pd.DataFrame(columns=SHORTLISTED_COLUMNS)
        ws = wb[SHORTLISTED_TAB_NAME]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame(columns=SHORTLISTED_COLUMNS)
    headers = [str(v) if v is not None else f"Col{i+1}" for i, v in enumerate(rows[0])]
    data = [list(row) for row in rows[1:] if any(v is not None for v in row)]
    df = pd.DataFrame(data, columns=headers)
    for col in ("Progress Report", "Award Status", "Assigned To", "Team Notes"):
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).replace("None", "")
    return df


def shortlist_solicitations(record_keys: list, actor: str) -> dict:
    """Atomically move records from Opportunities to Shortlisted in one upload.

    Keys are solicitation numbers, falling back to UiLink. Existing shortlisted
    records are skipped, and source rows are deleted only after their copy has
    been prepared in the same in-memory workbook.
    """
    from datetime import datetime, timezone

    wanted = {str(k).strip() for k in record_keys if str(k).strip()}
    if not wanted:
        return {"moved": 0, "skipped": 0, "missing": 0}

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()  # writes must start from the freshest Drive version
        wb, source, _ = _open_master()
        target = _ensure_tab(wb, SHORTLISTED_TAB_NAME, SHORTLISTED_COLUMNS)

        source_headers = {
            str(source.cell(1, c).value or "").strip(): c
            for c in range(1, source.max_column + 1)
        }
        target_headers = {
            str(target.cell(1, c).value or "").strip(): c
            for c in range(1, target.max_column + 1)
        }
        for column in SHORTLISTED_COLUMNS:
            if column not in target_headers:
                idx = target.max_column + 1
                target.cell(1, idx, column)
                target_headers[column] = idx

        def row_key(ws, row, headers):
            sol_col = headers.get("Solicitation Number")
            link_col = headers.get("UiLink")
            sol = str(ws.cell(row, sol_col).value or "").strip() if sol_col else ""
            link = str(ws.cell(row, link_col).value or "").strip() if link_col else ""
            return sol or link

        existing = {
            row_key(target, r, target_headers)
            for r in range(2, target.max_row + 1)
        }
        existing.discard("")
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        rows_to_delete = []
        moved = skipped = 0

        for row in range(2, source.max_row + 1):
            key = row_key(source, row, source_headers)
            if key not in wanted:
                continue
            if key in existing:
                skipped += 1
                rows_to_delete.append(row)
                continue

            new_row = target.max_row + 1
            for column in OUTPUT_COLUMNS:
                src_col = source_headers.get(column)
                if src_col:
                    target.cell(new_row, target_headers[column], source.cell(row, src_col).value)
            target.cell(new_row, target_headers["Shortlisted By"], actor)
            target.cell(new_row, target_headers["Shortlisted At"], now)
            target.cell(new_row, target_headers["Last Updated By"], actor)
            target.cell(new_row, target_headers["Last Updated At"], now)
            rows_to_delete.append(row)
            existing.add(key)
            moved += 1

        for row in sorted(rows_to_delete, reverse=True):
            source.delete_rows(row)

        if rows_to_delete:
            _apply_dropdown(target, target_headers)
            _save_and_upload(wb)
        else:
            wb.close()
        return {"moved": moved, "skipped": skipped, "missing": len(wanted) - moved - skipped}


def _worksheet_record_key(ws, row: int, headers: dict) -> str:
    """Return the stable solicitation identifier used by the scraper."""
    sol_col = headers.get("Solicitation Number")
    link_col = headers.get("UiLink")
    sol = str(ws.cell(row, sol_col).value or "").strip() if sol_col else ""
    link = str(ws.cell(row, link_col).value or "").strip() if link_col else ""
    return sol or link


@_serialized_workbook_write
def update_records_by_key(record_updates: dict, tab_name: str = SHORTLISTED_TAB_NAME) -> dict:
    """Apply changes to several records with one download and one upload.

    ``record_updates`` is ``{solicitation_number_or_link: {column: value}}``.
    Keys are resolved against the fresh workbook, so row insertions/deletions
    cannot redirect an edit to another solicitation.
    """
    wanted = {str(key).strip(): changes for key, changes in record_updates.items() if str(key).strip() and changes}
    if not wanted:
        return {"updated": 0, "missing": []}

    wb, _, _ = _open_master()
    ws = _ensure_tab(wb, tab_name, SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS)
    headers = {str(ws.cell(1, col).value or "").strip(): col for col in range(1, ws.max_column + 1)}
    found = set()
    changed = 0
    for row in range(2, ws.max_row + 1):
        key = _worksheet_record_key(ws, row, headers)
        changes = wanted.get(key)
        if not changes:
            continue
        found.add(key)
        for column, value in changes.items():
            col = headers.get(column)
            if col:
                ws.cell(row, col, value=value or None)
                changed += 1

    if changed:
        _save_and_upload(wb)
    else:
        wb.close()
    return {"updated": len(found), "missing": sorted(set(wanted) - found)}


@_serialized_workbook_write
def delete_records_by_key(record_keys: list, tab_name: str = SHORTLISTED_TAB_NAME) -> dict:
    """Delete records by stable key with one fresh-workbook upload."""
    wanted = {str(key).strip() for key in record_keys if str(key).strip()}
    if not wanted:
        return {"deleted": 0, "missing": []}
    wb, _, _ = _open_master()
    ws = _ensure_tab(wb, tab_name, SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS)
    headers = {str(ws.cell(1, col).value or "").strip(): col for col in range(1, ws.max_column + 1)}
    rows = []
    found = set()
    for row in range(2, ws.max_row + 1):
        key = _worksheet_record_key(ws, row, headers)
        if key in wanted:
            rows.append(row)
            found.add(key)
    for row in reversed(rows):
        ws.delete_rows(row)
    if rows:
        _save_and_upload(wb)
    else:
        wb.close()
    return {"deleted": len(rows), "missing": sorted(wanted - found)}


@_serialized_workbook_write
def delete_expired_rows(sheet_row_numbers: list, tab_name: str = SHEET_TAB_NAME) -> int:
    """Delete worksheet rows and upload the workbook once."""
    if not sheet_row_numbers:
        return 0
    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, main_ws, _ = _open_master()
        ws = main_ws if tab_name == SHEET_TAB_NAME else _ensure_tab(wb, tab_name, SHORTLISTED_COLUMNS)
        for row_num in sorted(sheet_row_numbers, reverse=True):
            ws.delete_rows(row_num)
        _save_and_upload(wb)
    log.info(f"Deleted {len(sheet_row_numbers)} row(s) from {tab_name}")
    return len(sheet_row_numbers)


@_serialized_workbook_write
def update_progress_reports(row_updates: dict, tab_name: str = SHEET_TAB_NAME) -> int:
    """
    Write Progress Report values back to the master sheet.

    row_updates: {sheet_row_number (int, 1-indexed, data starts at row 2): new_status (str)}
    Returns count of rows updated.
    """
    if not row_updates:
        return 0

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, _, _ = _open_master()
        ws = _ensure_tab(
            wb, tab_name,
            SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS,
        )

    # Locate Progress Report column
    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)}
    prog_col = headers.get("Progress Report")
    if not prog_col:
        raise ValueError("'Progress Report' column not found in master sheet")

    count = 0
    for sheet_row, new_status in row_updates.items():
        if 2 <= int(sheet_row) <= ws.max_row:
            ws.cell(row=int(sheet_row), column=prog_col, value=new_status or None)
            count += 1

    if count:
        _save_and_upload(wb)
        log.info(f"Updated {count} Progress Report value(s) in master sheet")

    return count


@_serialized_workbook_write
def update_assigned_to(row_updates: dict, tab_name: str = SHEET_TAB_NAME) -> int:
    """
    Write 'Assigned To' values back to the master sheet.

    row_updates: {sheet_row_number (int, 1-indexed, data starts at row 2): username (str)}
    Creates the column if it doesn't exist yet.
    Returns count of rows updated.
    """
    if not row_updates:
        return 0

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, _, _ = _open_master()
        ws = _ensure_tab(
            wb, tab_name,
            SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS,
        )
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    col = headers.get("Assigned To")

    if not col:
        col = ws.max_column + 1
        cell = ws.cell(row=1, column=col, value="Assigned To")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    count = 0
    for sheet_row, val in row_updates.items():
        if 2 <= int(sheet_row) <= ws.max_row:
            ws.cell(row=int(sheet_row), column=col, value=val or None)
            count += 1

    if count:
        _save_and_upload(wb)
        log.info(f"Updated {count} Assigned To value(s) in master sheet")

    return count


@_serialized_workbook_write
def update_award_status(row_updates: dict, tab_name: str = SHEET_TAB_NAME) -> int:
    """
    Write Award Status values back to the master sheet.
    Handles legacy sheets where the column may be called 'Award Date'.

    row_updates: {sheet_row_number (int): status_string}
    Returns count of rows updated.
    """
    if not row_updates:
        return 0

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, _, _ = _open_master()
        ws = _ensure_tab(
            wb, tab_name,
            SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS,
        )
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}

    # Support legacy "Award Date" header alongside new "Award Status"
    col = headers.get("Award Status") or headers.get("Award Date")
    if not col:
        col = ws.max_column + 1
        cell = ws.cell(row=1, column=col, value="Award Status")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    count = 0
    for sheet_row, val in row_updates.items():
        if 2 <= int(sheet_row) <= ws.max_row:
            ws.cell(row=int(sheet_row), column=col, value=val or None)
            count += 1

    if count:
        _save_and_upload(wb)
        log.info(f"Updated {count} Award Status value(s) in master sheet")

    return count


@_serialized_workbook_write
def update_shortlisted_record(sheet_row: int, updates: dict,
                               tab_name: str = SHORTLISTED_TAB_NAME) -> int:
    """Update multiple columns in one row with a single Drive upload.

    updates: {column_name: new_value}
    Returns 1 if anything was written, 0 otherwise.
    Used by the detail popup so saving 3 fields costs only 1 round-trip.
    """
    if not updates or sheet_row < 2:
        return 0

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, _, _ = _open_master()
        ws = _ensure_tab(
            wb, tab_name,
            SHORTLISTED_COLUMNS if tab_name == SHORTLISTED_TAB_NAME else OUTPUT_COLUMNS,
        )

    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    if sheet_row > ws.max_row:
        return 0

    changed = 0
    for col_name, new_val in updates.items():
        col_idx = headers.get(col_name)
        if col_idx:
            ws.cell(row=sheet_row, column=col_idx, value=new_val or None)
            changed += 1

    if changed:
        _save_and_upload(wb)
        log.info(f"update_shortlisted_record: row {sheet_row}, {changed} column(s) written")
    return 1 if changed else 0


# ─────────────────────────────────────────────
# Date helper (single value, no pandas)
# ─────────────────────────────────────────────

def _parse_date(val):
    """Parse a cell value to a date object. Returns None if unparseable."""
    from datetime import date, datetime as dt
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, dt):
        return val
    if isinstance(val, dt):
        return val.date()
    s = str(val).strip()[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────
# Urgent tab
# ─────────────────────────────────────────────

def _ensure_urgent_tab(wb):
    """Return the Urgent worksheet, creating it with headers if needed."""
    if URGENT_TAB_NAME not in wb.sheetnames:
        ws = wb.create_sheet(URGENT_TAB_NAME)
        fill = PatternFill("solid", fgColor="C00000")
        font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        widths = [20, 40, 35, 18, 18, 20, 35, 40, 20, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        log.info(f"Created '{URGENT_TAB_NAME}' sheet tab")
    return wb[URGENT_TAB_NAME]


@_serialized_workbook_write
def refresh_urgent_tab() -> tuple:
    """Rebuild Urgent as a read-only projection of the Shortlisted queue."""
    from datetime import date, timedelta
    today        = date.today()
    urgent_limit = today + timedelta(days=10)

    with _WORKBOOK_LOCK:
        _invalidate_raw_cache()
        wb, _, _ = _open_master()
        source = _ensure_tab(wb, SHORTLISTED_TAB_NAME, SHORTLISTED_COLUMNS)
        urgent = _ensure_urgent_tab(wb)
        removed = max(urgent.max_row - 1, 0)
        if removed:
            urgent.delete_rows(2, removed)

        headers = {
            str(source.cell(1, c).value or "").strip(): c
            for c in range(1, source.max_column + 1)
        }
        due_col = headers.get("Due Date")
        added = 0
        for row in range(2, source.max_row + 1):
            due = _parse_date(source.cell(row, due_col).value) if due_col else None
            if not due or not (today <= due < urgent_limit):
                continue
            added += 1
            for column_index, column in enumerate(OUTPUT_COLUMNS, 1):
                source_col = headers.get(column)
                urgent.cell(added + 1, column_index, source.cell(row, source_col).value if source_col else None)

        _save_and_upload(wb)
    log.info(f"Urgent tab rebuilt from Shortlisted: {removed} removed, {added} added")
    return removed, added


def read_urgent_tab():
    """Return the Urgent sheet tab as a DataFrame."""
    import pandas as pd
    raw = _get_master_bytes()
    wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if URGENT_TAB_NAME not in wb.sheetnames:
        wb.close()
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    ws = wb[URGENT_TAB_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    headers = [str(v) if v is not None else f"Col{i}" for i, v in enumerate(rows[0])]
    data    = [list(r) for r in rows[1:] if any(v is not None for v in r)]
    df = pd.DataFrame(data, columns=headers)
    if "Progress Report" in df.columns:
        df["Progress Report"] = df["Progress Report"].fillna("").astype(str).replace("None", "")
    return df


# ─────────────────────────────────────────────
# Overdue cleanup (main sheet)
# ─────────────────────────────────────────────

@_serialized_workbook_write
def cleanup_overdue_rows() -> int:
    """
    Delete rows from the main sheet where:
      - due_date is more than 5 days in the past
      - Progress Report is empty
    Returns number of rows deleted.
    """
    from datetime import date, timedelta
    cutoff = date.today() - timedelta(days=5)

    wb, ws, _ = _open_master()
    hdrs = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    due_col  = hdrs.get("Due Date")
    prog_col = hdrs.get("Progress Report")

    to_delete = []
    for r in range(2, ws.max_row + 1):
        due  = _parse_date(ws.cell(r, due_col).value) if due_col  else None
        prog = str(ws.cell(r, prog_col).value or "").strip() if prog_col else ""
        if due and due < cutoff and not prog:
            to_delete.append(r)

    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r)

    if to_delete:
        _save_and_upload(wb)
        log.info(f"Cleaned up {len(to_delete)} overdue empty-progress rows")

    return len(to_delete)
