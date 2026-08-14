import io
import sys
import types
import unittest
from unittest.mock import patch

import openpyxl

# The connector's network functions are mocked in these unit tests; a tiny stub
# keeps the test independent of whether the runtime bundles requests.
sys.modules.setdefault("requests", types.SimpleNamespace())
import google_connector as connector


def workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = connector.SHEET_TAB_NAME
    sheet.append(connector.OUTPUT_COLUMNS)
    sheet.append([
        "SOL-001", "Generator", "Army", "2026-08-01", "2026-09-01",
        "Solicitation", "SDVOSB", "https://sam.gov/1", "", "", "",
    ])
    sheet.append([
        "SOL-002", "Roof Repair", "Navy", "2026-08-02", "2026-09-02",
        "Solicitation", "SBA", "https://sam.gov/2", "", "", "",
    ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ShortlistTests(unittest.TestCase):
    def setUp(self):
        connector._invalidate_raw_cache()
        self.remote_bytes = workbook_bytes()

    def download(self, _file_id):
        return self.remote_bytes

    def upload(self, _file_id, content, _mime_type):
        self.remote_bytes = content
        return {"id": "test"}

    def test_move_is_single_upload_and_preserves_other_rows(self):
        with patch.object(connector, "download_drive_file", side_effect=self.download), patch.object(
            connector, "upload_drive_file", side_effect=self.upload
        ) as upload:
            result = connector.shortlist_solicitations(["SOL-001"], "tobi")

        self.assertEqual(result, {"moved": 1, "skipped": 0, "missing": 0})
        self.assertEqual(upload.call_count, 1)
        workbook = openpyxl.load_workbook(io.BytesIO(self.remote_bytes), data_only=True)
        opportunities = workbook[connector.SHEET_TAB_NAME]
        shortlisted = workbook[connector.SHORTLISTED_TAB_NAME]
        self.assertEqual(opportunities.cell(2, 1).value, "SOL-002")
        self.assertEqual(shortlisted.cell(2, 1).value, "SOL-001")
        headers = [cell.value for cell in shortlisted[1]]
        self.assertEqual(
            shortlisted.cell(2, headers.index("Shortlisted By") + 1).value,
            "tobi",
        )

    def test_deduplication_checks_shortlisted_and_opportunities(self):
        with patch.object(connector, "download_drive_file", side_effect=self.download), patch.object(
            connector, "upload_drive_file", side_effect=self.upload
        ):
            connector.shortlist_solicitations(["SOL-001"], "tobi")
            connector._invalidate_raw_cache()
            solicitations, links = connector.get_existing_dedup_keys()

        self.assertEqual(solicitations, {"SOL-001", "SOL-002"})
        self.assertEqual(links, {"https://sam.gov/1", "https://sam.gov/2"})


if __name__ == "__main__":
    unittest.main()
